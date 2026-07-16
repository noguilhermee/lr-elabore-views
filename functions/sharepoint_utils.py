# ==============================
# 📦 Bibliotecas padrão (Python)
# ==============================
import os
import sys
import json
import csv
import time
import re
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from zoneinfo import ZoneInfo

# ==============================
# 🌐 Requisições e dados
# ==============================
import requests
import pandas as pd

# =========================
# Excel
# =========================
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ==============================
# 🗄️ Banco de dados
# ==============================
import psycopg2
from psycopg2.extras import execute_batch

# ==============================
# 🔐 Configuração e ambiente
# ==============================
from dotenv import load_dotenv
from venv import logger


# ==============================
# ⚙️ Inicialização
# ==============================
inicio = time.time()

arquivo_env = load_dotenv()
print(f"Variável environmental carregada: {arquivo_env}")
print("Pacotes carregados!")

# Diretório atual do notebook
diretorio = os.getcwd()

# Caminho do .env
env_path = os.path.join(diretorio, '.env')

# Carregar variáveis de ambiente do .env
load_dotenv(env_path, override=True)

# Trazer as variáveis de ambiente
tenant_id = os.getenv('TENANT_ID')
client_id = os.getenv('CLIENT_ID')
client_secret = os.getenv('CLIENT_SECRET')
site_id = os.getenv('SITE_ID')
list_id = os.getenv('TAB_PARCELA')

# Defina seu fuso horário (ajuste conforme sua localização)
LOCAL_TZ = ZoneInfo("America/Sao_Paulo")  # UTC-3 (horário de Brasília)

CHECKPOINT_FILE = 'sharepoint_checkpoint.json'


#
# Formatar excel
#

def formatar_excel(arquivo_excel):
    wb = load_workbook(arquivo_excel)

    fonte_padrao    = Font(name="Aptos", size=10)
    fonte_cabecalho = Font(name="Aptos", size=10, bold=True)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        # Altura padrão de todas as linhas de dados (muito mais rápido que row_dimensions em loop)
        ws.sheet_format.defaultRowHeight = 18
        ws.sheet_format.customHeight = True

        # Primeira linha: altura e cabeçalho
        ws.row_dimensions[1].height = 30
        for cell in ws[1]:
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_centro

        # Largura e fonte por coluna inteira (evita iter_rows célula a célula)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in column_cells:
                # Aplicar fonte e alinhamento (só nas linhas de dados, cabeçalho já foi feito)
                if cell.row > 1:
                    cell.font = fonte_padrao
                    cell.alignment = alinhamento_centro

                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

        # Filtro
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(arquivo_excel)

def formatar_excel_antigo(arquivo_excel):
    """
    Formata todas as abas do arquivo Excel:
    - 1ª linha com altura 30
    - demais linhas com altura 18
    - fonte Aptos tamanho 10
    - texto centralizado horizontal e verticalmente
    - largura automática das colunas
    - filtro na primeira linha
    - congelamento da primeira linha
    """

    wb = load_workbook(arquivo_excel)

    fonte_padrao = Font(name="Aptos", size=10)
    fonte_cabecalho = Font(name="Aptos", size=10, bold=True)

    alinhamento_centro = Alignment(
        horizontal="center",
        vertical="center"
    )

    for ws in wb.worksheets:

        # Ocultar linhas de grade da planilha
        ws.sheet_view.showGridLines = False

        # Congelar primeira linha
        ws.freeze_panes = "A2"

        # Altura da primeira linha
        ws.row_dimensions[1].height = 30

        # Aplicar formatação
        for row in ws.iter_rows():
            for cell in row:
                cell.font = fonte_padrao
                cell.alignment = alinhamento_centro

        # Cabeçalho em negrito
        for cell in ws[1]:
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_centro

        # Altura das demais linhas
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 18

        # Ajustar largura das colunas
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for cell in column_cells:
                if cell.value is not None:
                    valor = str(cell.value)
                    max_length = max(max_length, len(valor))

            largura = min(max(max_length + 2, 12), 45)
            ws.column_dimensions[col_letter].width = largura

        # Ativar filtro
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(arquivo_excel)

def exportar_excel_formatado(df, arquivo_excel, nome_aba="Dados"):
    """
    Exporta um DataFrame para Excel e aplica a formatação padrão.
    """

    arquivo_excel = Path(arquivo_excel)
    arquivo_excel.parent.mkdir(parents=True, exist_ok=True)

    df.to_excel(
        arquivo_excel,
        sheet_name=nome_aba,
        index=False
    )

    formatar_excel(arquivo_excel)

    print(f"Exportado e formatado: {arquivo_excel}")

# ========== FUNÇÕES DE CHECKPOINT ==========
def load_checkpoint(list_name: str) -> Optional[str]:
    if not os.path.exists(CHECKPOINT_FILE):
        return None
    with open(CHECKPOINT_FILE, 'r') as f:
        checkpoints = json.load(f)
    return checkpoints.get(list_name)

def save_checkpoint(list_name: str, timestamp: str):
    checkpoints = {}
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, 'r') as f:
            checkpoints = json.load(f)
    checkpoints[list_name] = timestamp
    with open(CHECKPOINT_FILE, 'w') as f:
        json.dump(checkpoints, f, indent=2)
    print(f"✓ Checkpoint salvo: {list_name} -> {timestamp}")

def get_incremental_date(list_name: str, days_back: int = 1) -> str:
    checkpoint = load_checkpoint(list_name)
    if checkpoint:
        # Se o checkpoint existir
        print(f"📌 Usando checkpoint: {checkpoint}")
        return checkpoint
    
    else:
        # Se for a primeira vez
        fallback_date = (datetime.now(LOCAL_TZ) - timedelta(days=days_back)).isoformat()
        print(f"⚠️  Sem checkpoint. Usando fallback: últimos {days_back} dia(s)")
        return fallback_date

# ========== AUTENTICAÇÃO ==========
def get_access_token():
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://graph.microsoft.com/.default'
    }
    try:
        # Adicionar timeout de 1 minuto para a requisicao de token
        token_response = requests.post(token_url, data=data, timeout=60)
        
        token_response.raise_for_status() # Levanta HTTPError para códigos de status de erro
        access_token = token_response.json().get('access_token')
        
        print('✓ Token obtido:', access_token[:40] + '...')
        return access_token
    
    except requests.exceptions.Timeout:
        logger.error("❌ Erro de Timeout ao obter token do SharePoint.")

        # Adicionado print para feedback
        print("❌ Erro de Timeout ao obter token do SharePoint.")
        return None
    
    except requests.exceptions.RequestException as e:
        logger.error(f"❌ Erro ao obter token do SharePoint: {e}")

        # Adicionado print para feedback
        print(f"❌ Erro ao obter token do SharePoint: {e}")
        return None

# ========== EXTRAÇÃO INCREMENTAL ==========
def get_sharepoint_data_incremental(
    list_id: str, 
    list_name: str, 
    access_token: str,
    incremental: bool = True,
    days_back: int = 1
) -> List[Dict]:
    start = time.time()

    if not access_token: 
        # Adicionar verificação se o token foi obtido
        logger.error(f"❌ Não foi possível extrair dados para {list_name}: Token de acesso inválido.")
        print(f"❌ Não foi possível extrair dados para {list_name}: Token de acesso inválido.") # Adicionado print
        return []
        
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    base_url = (
        f'https://graph.microsoft.com/v1.0/sites/laborruralserv.sharepoint.com,{site_id}'
        f'/lists/{list_id}/items'
    )

    params = {
        '$expand': 'fields',
        '$top': 5000
    }

    if incremental:
        reference_date = get_incremental_date(list_name, days_back)
        params['$filter'] = f"fields/Modified ge '{reference_date}'"
        print(f"🔍 Filtro incremental: Modified >= {reference_date}")
    else:
        print(f"📥 Carga completa (sem filtro)")

    all_items = []
    page_count = 0

    try:
        response = requests.get(base_url, headers=headers, params=params, timeout=60)
        logger.debug(f"🌐 URL gerada para {list_name}: {response.url}")
        print(f"🌐 URL gerada para {list_name}: {response.url}\n")

        while True:
            page_count += 1
            print(f"Página {page_count:<2}...", end=' ')

            response.raise_for_status() # Levanta HTTPError para códigos de status de erro (4xx ou 5xx) 
            data = response.json()
            items = data.get('value', [])
            all_items.extend(items)
            print(f"✓ {len(items):<2} itens | Total: {len(all_items):<2}")
            
            next_link = data.get('@odata.nextLink')
            if not next_link:
                break
            response = requests.get(next_link, headers=headers, timeout=60)

    except requests.exceptions.Timeout:
        logger.error(f"❌ Erro de Timeout ao extrair dados da lista '{list_name}' do SharePoint. A requisição demorou mais que o esperado.")
        print(f"❌ Erro de Timeout ao extrair dados da lista '{list_name}' do SharePoint. A requisição demorou mais que o esperado.") # Adicionado print
        return []
    
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 400 and incremental:
            logger.warning(f"❌ Erro 400 ao extrair dados da lista '{list_name}' com filtro incremental. Resposta: {e.response.text}")
            print(f"❌ Erro 400 ao extrair dados da lista '{list_name}' com filtro incremental. Resposta: {e.response.text}") # Adicionado print
            logger.warning("⚠️  Tentando carga completa como fallback...")
            print("⚠️  Tentando carga completa como fallback...") # Adicionado print
            return get_sharepoint_data_incremental(
                list_id, list_name, access_token, 
                incremental=False, days_back=days_back
            )
        else:
            logger.error(f"❌ Erro HTTP ao extrair dados da lista '{list_name}': {e}")
            print(f"❌ Erro HTTP ao extrair dados da lista '{list_name}': {e}") # Adicionado print
            logger.error(f"Resposta do SharePoint: {e.response.text}")
            print(f"Resposta do SharePoint: {e.response.text}") # Adicionado print
        return []
    
    except requests.exceptions.RequestException as e:
        logger.error(f"❌ Erro de conexão ao extrair dados da lista '{list_name}': {e}")
        print(f"❌ Erro de conexão ao extrair dados da lista '{list_name}': {e}") # Adicionado print
        return []
    
    except Exception as e:
        logger.error(f"❌ Erro inesperado ao extrair dados da lista '{list_name}': {e}")
        print(f"❌ Erro inesperado ao extrair dados da lista '{list_name}': {e}") # Adicionado print
        logger.exception("Detalhes do erro:")
        return []
    
    logger.info(f"✓ Lista {list_name}: {len(all_items)} itens extraídos com sucesso.")
    print(f"✓ Lista {list_name}: {len(all_items)} itens extraídos")

    if incremental and all_items:
        current_timestamp = datetime.now(LOCAL_TZ).isoformat()
        save_checkpoint(list_name, current_timestamp)

    return all_items

# ========== CONVERSÕES DE TIPO ==========
def apply_data_conversions(df: pd.DataFrame, list_type: str):
    """
    Aplica conversões de tipo específicas para cada tipo de lista
    """
    if list_type == 'fazenda':
        numeric_columns = ['idProdutor', 'idAgroindustria', 'Excluido', 'cep', 'numero', 'editadoPor', 'criadoPor']
        date_columns = ['dteEntradaInicial', 'Criado', 'Modificado','dteSaida']
        
    elif list_type == 'parcelas':
        numeric_columns = ['ID', 'idInventario', 'numeroParcela', 'valor', 'Excluido', 'Criado por', 'Modificado por']
        date_columns = ['Criado', 'Modificado', 'previsaoPagamento', 'dataPagamento']
        
    elif list_type == 'inventario':
        numeric_columns = ['ID', 'idFazenda', 'quantidade', 'valorAquisicao', 'vidaUtil', 'Excluido', 'Criado por', 'Modificado por']
        date_columns = ['dataAquisicao', 'dataFim', 'Criado', 'Modificado']
    
    elif list_type == 'componentesrendabruta':
        numeric_columns = [
            'ID', 'idFazenda', 'leiteVendido', 'leiteConsumido', 'precoLeite', 
            'vendaAnimais', 'qtdeVendaVolumoso', 'qtdeVendaConcentrado', 
            'outrasReceitas', 'derivadosVenda', 'derivadosPreco', 
            'bonificacaoPreco', 'divisaodesobrasPreco', 'parcela', 
            'emprestimosRecebidos', 'parcelaEmprestimosRecebidos'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']
    
    elif list_type == 'componentescusto':
        numeric_columns = [
            'ID', 'idFazenda', 'gastoSucedaneo', 'gastoMedicamentosVacinas', 
            'gastoHormonios', 'gastoMaterialOrdenha', 'gastoEnergiaEletrica', 
            'gastoCombustivel', 'gastoReproducao', 'gastoImpostoTaxas', 
            'gastoAdministrativo', 'gastoReparosConsertos', 'gastoAssistenciaTecnica', 
            'gastoCamaAreia', 'gastoAcessoriosDespesasGerais', 'gastoArrendamento', 
            'consumoLeiteBezerro', 'consumoMaoObraFamiliar', 'consumoMaoObraContratada', 
            'consumoLeiteDescartado', 'compra_Animais', 'compra_Terras', 
            'parcelaReposicaoCama', 'parcelaCompraAnimais', 'parcelaCompraTerra', 
            'parcelaSucedaneo', 'parcelaMaterialOrdenha', 'parcelaReproducao', 
            'parcelaHormonios', 'parcelaMedicamentosVacinas', 'parcelaImpostoTaxas', 
            'parcelaAssistenciaTecnica', 'parcelaArrendamento', 'parcelaReparosConsertos', 
            'parcelaAdministrativo', 'parcelaAcessoriosDespesasGerais', 
            'gastoEmprestimosJurosPagos', 'parcelaEmprestimosJurosPagos'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']
    
    # ========== ALIMENTAÇÃO ==========
    elif list_type == 'alimentacao':
        numeric_columns = [
            'ID', 'idFazenda', 'quantidadeComprada', 'quantidadeConsumida', 
            'valorUnitario', 'Parcelamento', 'Collection', 'Excluido', 
            'Update', 'id_item'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']

    # ========== ÁREA ==========
    elif list_type == 'area':
        numeric_columns = [
            'ID', 'idFazenda', 'tamanhoArea', 'precoTerraNua', 
            'Excluido', 'Collection', 'Update', 'arrendamento'
        ]
        date_columns = ['dataInicio', 'dataFim', 'Criado', 'Modificado']

    # ========== CONSULTOR ==========
    elif list_type == 'consultor':
        numeric_columns = [
            'ID', 'idConsultoria', 'Excluido', 'idGestor', 
            'editadoPor', 'criadoPor'
        ]
        date_columns = ['dteNascConsultor', 'dataInicio', 'dataFim', 'Criado', 'Modificado']

    # ========== CULTURAS ==========
    elif list_type == 'culturas':
        numeric_columns = [
            'ID', 'idFazenda', 'idArea', 'Collection', 'Excluido', 
            'areaha', 'Update', 'idEspecificacaoCultura'
        ]
        date_columns = ['dataInicio', 'dataFim', 'Plantio', 'Criado', 'Modificado']
        boolean_columns = ['finalizado']

    # ========== CUSTO FORRAGEIRA ==========
    elif list_type == 'custoforrageira':
        numeric_columns = [
            'ID', 'idFazenda', 'idCultura', 'qtdeConsumida', 'qtdeComprada', 
            'valorUnitario', 'n', 'p', 'k', 'numeroEspecies', 'Collection', 
            'Excluido', 'Update', 'Parcelas', 'id_item', 'idArea'
        ]
        date_columns = ['Criado', 'Modificado']

    # ========== ENERGIA COMBUSTÍVEL ==========
    elif list_type == 'energiacombustivel':
        numeric_columns = [
            'ID', 'qtdeEnergia', 'valorEnergia', 'qtdeAlcoolgasolina', 
            'valorAlcoolgasolina', 'qtdeDiesel', 'valorDiesel', 'idFazenda', 
            'qtdeEnergiaSolar', 'valorEnergiaSolar'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']

    # ========== ENTRADA ESTOQUE ==========
    elif list_type == 'entradaestoque':
        numeric_columns = [
            'ID', 'idRelacionamento', 'idItem', 'idFazenda', 
            'qntComprada', 'valorUnitario', 'qntConsumida', 'Excluido'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']

    # ========== MÃO DE OBRA ==========
    elif list_type == 'maodeobra':
        numeric_columns = [
            'ID', 'idFazenda', 'qtdeCustoMaodeObra', 'valorCustoMaodeObra',
            'familiarQtd', 'familiarValortotal', 'contratataordenhadorQtd',
            'contratataordenhadorValortotal', 'ctdServicosgeraisQtd',
            'ctdServicosgeraisValor', 'ctdFolguistaQtde', 'ctdFolguistaValor',
            'ctdTratoristaQtde', 'ctdTratoristaValor', 'encargostrabalhistasValor',
            'demaisdespesasValor'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']
    
    # ========== PRODUÇÃO ==========
    elif list_type == 'producao':
        numeric_columns = [
            'ID', 'idItem', 'areaProduzida', 'producao', 
            'idFazenda', 'idCultura', 'Excluido'
        ]
        date_columns = ['dataColheita', 'Criado', 'Modificado']

    # ========== PRODUTOR ==========
    elif list_type == 'produtor':
        numeric_columns = [
            'ID', 'idProdutor', 'celProdutor', 'cepProdutor', 
            'Excluido', 'editadoPor', 'criadoPor'
        ]
        date_columns = [
            'dteNascProdutor', 'dataInicio', 'dataFim', 
            'dataretorno', 'Criado', 'Modificado'
        ]
        boolean_columns = ['inativado']

    # ========== QUALIDADE LEITE ==========
    elif list_type == 'qualidadeleite':
        numeric_columns = [
            'ID', 'idFazenda', 'CCS', 'CPP', 'Gordura', 'Proteina'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']

    # ========== REBANHO ==========
    elif list_type == 'rebanho':
        numeric_columns = [
            'ID', 'idFazenda', 'qtdeVacasEmLactacao', 'qtdeVacasSecas',
            'qtdeAnimaisAleitamento', 'qtdeAnimaisRecria', 'qtdeOutrasCategorias',
            'vidautilOutrasCategorias', 'valorunitVacasEmLactacao', 'valorunitVacasSecas',
            'valorunitAnimaisAleitamento', 'valorunitAnimaisRecria', 'valorunitOutrasCategorias',
            'qtdeMacho', 'valorunitMacho', 'compraVacasEmLactacao', 'compraVacasSecas',
            'compraAnimaisAleitamento', 'compraAnimaisRecria', 'compraMacho',
            'compraOutrasCategorias', 'vendaVacasEmLactacao', 'vendaVacasSecas',
            'vendaAnimaisAleitamento', 'vendaAnimaisRecria', 'vendaMacho',
            'vendaOutrasCategorias', 'morteVacasEmLactacao', 'morteVacasSecas',
            'morteAnimaisAleitamento', 'morteAnimaisRecria', 'morteMacho',
            'morteOutrasCategorias'
        ]
        date_columns = ['Criado', 'Modificado','mesReferencia']

    # ========== RECRIA ==========
    elif list_type == 'recria':
        numeric_columns = [
            'ID', 'idFazenda', 'percentualVacasCompostBarn', 'percentualVacasFreeStall',
            'percentualVacasSemConfinado', 'percentualVacasPasto',
            'percentualVacasConfinadoSemEstrutura', 'Excluido', 'Collection', 'Update'
        ]
        date_columns = ['dataInicio', 'dataFim', 'Criado', 'Modificado']

    # ========== SAÍDA ESTOQUE ==========
    elif list_type == 'saidaestoque':
        numeric_columns = [
            'ID', 'idRelacionamento', 'idItem', 'idFazenda', 
            'qntConsumida', 'valorUnitario', 'idEstoque', 'Excluido'
        ]
        date_columns = ['mesReferencia', 'Criado', 'Modificado']

    # ========== SISTEMA PRODUÇÃO ==========
    elif list_type == 'sistemaproducao':
        numeric_columns = [
            'ID', 'idFazenda', 'percentualVacasCompostBarn', 'percentualVacasFreeStall',
            'percentualVacasSemConfinado', 'percentualVacasConfinadoSemEstrutura',
            'percentualVacasPasto', 'Excluido', 'Collection', 'Update'
        ]
        date_columns = ['dataInicio', 'dataFim', 'Criado', 'Modificado']

    # ========== VÍNCULOS ==========
    elif list_type == 'vinculos':
        numeric_columns = [
            'ID', 'IDFazenda', 'IDConsultor', 'IDProdutor'
        ]
        date_columns = [
            'dataAprovacao', 'dataReprovacao', 'dataInativacao', 
            'Criado', 'Modificado'
        ]

    # ========== AGROINDÚSTRIA ==========
    elif list_type == 'agroindustria':
        numeric_columns = [
            'ID', 'cepSede', 'numero', 'Excluido', 'editadoPor', 'criadoPor'
        ]
        date_columns = ['dataInicio', 'dataFim', 'Criado', 'Modificado']
        
    # ========== NOVO: ITENS PATRIMÔNIO ==========
    elif list_type == 'itenspatrimonio':
        numeric_columns = [
            'ID'
        ]
        date_columns = ['Criado', 'Modificado']
        
    else:
        numeric_columns = []
        date_columns = []
        boolean_columns = []
    
    # Aplicar conversões numéricas
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Aplicar conversões de data
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Aplicar conversões booleanas (se houver)
    if 'boolean_columns' in locals():
        for col in boolean_columns:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: True if x in [True, 1, '1', 'Sim', 'Yes', 'true'] 
                    else False if x in [False, 0, '0', 'Não', 'No', 'false'] 
                    else None
    )

# ========== PROCESS DATAFRAME (GENÉRICO) ==========
def process_dataframe(items: List[Dict], column_mapping: Dict[str, str], target_columns: List[str], list_type: str) -> pd.DataFrame:
    """
    Processa os itens do SharePoint em DataFrame com mapeamento de colunas
    """
    data_list = []
    
    for item in items:
        fields = item.get('fields', {})
        row = {}
        
        # Mapear campos existentes
        for sharepoint_col, df_col in column_mapping.items():
            if sharepoint_col in fields:
                row[df_col] = fields[sharepoint_col]
            elif sharepoint_col in item:  
                # Para campos como Created, Modified
                row[df_col] = item[sharepoint_col]
        
        data_list.append(row)
    
    # Criar DataFrame
    df = pd.DataFrame(data_list)
    
    # Garantir que todas as colunas existam
    for col in target_columns:
        if col not in df.columns:
            df[col] = None
    
    # Reordenar colunas
    df = df.reindex(columns=target_columns)
    
    # Conversões de tipo baseadas no tipo de lista
    # apply_data_conversions(df, list_type)
    
    return df

# ========== CONVERSÃO PARA DATAFRAMES ==========
def fazenda_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    column_mapping = {
        'id': 'id',
        'Excluido': 'Excluido',
        'idProdutor': 'idProdutor',
        'idConsultor': 'idConsultor',
        'aprovacao': 'aprovacao',
        'inscricaoEstadual': 'inscricaoEstadual',
        'idAgroindustria': 'idAgroindustria',
        'codAgroindustria': 'codAgroindustria',
        'cidadeFazenda': 'cidadeFazenda',
        'ufFazenda': 'ufFazenda',
        'latitudeFazenda': 'latitudeFazenda',
        'longitudeFazenda': 'longitudeFazenda',
        'altitudeFazenda': 'altitudeFazenda',
        'dteEntradaInicial': 'dteEntradaInicial',
        'dteSaida':'dteSaida',
        'regiaoLeiteira': 'regiaoLeiteira',
        'Status': 'Status',
        'cep': 'cep',
        'logradouro': 'logradouro',
        'bairro': 'bairro',
        'numero': 'numero',
        'editadoPor': 'editadoPor',
        'criadoPor': 'criadoPor',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'nomeFazenda': 'nomeFazenda'
    }
    
    target_columns = [
        'id', 'Excluido', 'idProdutor', 'idConsultor', 'aprovacao', 'inscricaoEstadual',
        'idAgroindustria', 'nomeFazenda', 'codAgroindustria', 'cidadeFazenda', 
        'ufFazenda', 'latitudeFazenda', 'longitudeFazenda', 'altitudeFazenda', 
        'dteEntradaInicial', 'dteSaida', 'regiaoLeiteira', 'Status', 'cep', 'logradouro', 
        'bairro', 'numero', 'editadoPor', 'criadoPor', 'Criado', 'Modificado'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'fazenda')

def itenspatrimonio_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_itenspatrimonio em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'categoria': 'categoria',
        'patrimonio': 'patrimonio',
        'Created': 'Criado',
        'Modified': 'Modificado'
    }
    
    target_columns = [
        'ID', 'categoria', 'patrimonio',
        'Criado', 'Modificado'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'itenspatrimonio')

def parcelas_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    column_mapping = {
        'id': 'ID',
        'idInventario': 'idInventario',
        'Item': 'Item',
        'nomeTela': 'nomeTela',
        'Excluido': 'Excluido',
        'numeroParcela': 'numeroParcela',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'valor': 'valor',
        'previsaoPagamento': 'previsaoPagamento',
        'dataPagamento': 'dataPagamento',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idInventario', 'Item', 'id_Item', 'nomeTela', 'Excluido',
        'numeroParcela', 'Criado', 'Modificado', 'valor', 'previsaoPagamento', 
        'dataPagamento', 'Criado por', 'Modificado por'
    ]
    
    df = process_dataframe(items, column_mapping, target_columns, 'parcelas')
    # df['dataPagamento'] = df['dataPagamento'] - pd.Timedelta(hours=3)
    # df['previsaoPagamento'] = df['previsaoPagamento'] - pd.Timedelta(hours=3)
    df['id_Item'] = None
    
    return df

def inventario_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'dataAquisicao': 'dataAquisicao',
        'classificacao': 'classificacao',
        'tipoItem': 'tipoItem',
        'nomeItem': 'nomeItem',
        'valorAquisicao': 'valorAquisicao',
        'vidaUtil': 'vidaUtil',
        'quantidade': 'quantidade',
        'Excluido': 'Excluido',
        'Parcelamento': 'Parcelamento',
        'AuthorLookupId': 'Criado por',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'dataAquisicao', 'classificacao', 'tipoItem', 'nomeItem',
        'valorAquisicao', 'vidaUtil', 'dataFim', 'quantidade', 'Excluido', 'Parcelamento',
        'Criado por', 'Criado', 'Modificado', 'Modificado por'
    ]
    
    df = process_dataframe(items, column_mapping, target_columns, 'inventario')
    df['dataFim'] = None
    
    return df

# ========== NOVO: COMPONENTES RENDA BRUTA ==========
def componentesrendabruta_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_componentesrendabruta em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'mesReferencia': 'mesReferencia',
        'leiteVendido': 'leiteVendido',
        'leiteConsumido': 'leiteConsumido',
        'precoLeite': 'precoLeite',
        'vendaAnimais': 'vendaAnimais',
        'qtdeVendaVolumoso': 'qtdeVendaVolumoso',
        'qtdeVendaConcentrado': 'qtdeVendaConcentrado',
        'outrasReceitas': 'outrasReceitas',
        'Observacao': 'Observacao',
        'derivadosVenda': 'derivadosVenda',
        'derivadosPreco': 'derivadosPreco',
        'bonificacaoPreco': 'bonificacaoPreco',
        'divisaodesobrasPreco': 'divisaodesobrasPreco',
        'observacaoLeite': 'observacaoLeite',
        'observacaoDerivados': 'observacaoDerivados',
        'observacaoBonificacao': 'observacaoBonificacao',
        'observacaoDivisao': 'observacaoDivisao',
        'observacaoVendaanimais': 'observacaoVendaanimais',
        'observacaoOutrasreceitas': 'observacaoOutrasreceitas',
        'itemdoestoqueVolumoso': 'itemdoestoqueVolumoso',
        'itemdoestoqueConcentrado': 'itemdoestoqueConcentrado',
        'parcela': 'parcela',
        'aux': 'aux',
        'emprestimosRecebidos': 'emprestimosRecebidos',
        'observacaoEmprestimosRecebidos': 'observacaoEmprestimosRecebidos',
        'parcelaEmprestimosRecebidos': 'parcelaEmprestimosRecebidos',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'mesReferencia', 'leiteVendido', 'leiteConsumido',
        'precoLeite', 'vendaAnimais', 'qtdeVendaVolumoso', 'qtdeVendaConcentrado',
        'outrasReceitas', 'Observacao', 'derivadosVenda', 'derivadosPreco',
        'bonificacaoPreco', 'divisaodesobrasPreco', 'observacaoLeite',
        'observacaoDerivados', 'observacaoBonificacao', 'observacaoDivisao',
        'observacaoVendaanimais', 'observacaoOutrasreceitas', 'itemdoestoqueVolumoso',
        'itemdoestoqueConcentrado', 'parcela', 'aux', 'emprestimosRecebidos',
        'observacaoEmprestimosRecebidos', 'parcelaEmprestimosRecebidos',
        'Criado', 'Modificado', 'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'componentesrendabruta')

def componentescusto_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_componentescusto em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'gastoSucedaneo': 'gastoSucedaneo',
        'gastoMedicamentosVacinas': 'gastoMedicamentosVacinas',
        'gastoHormonios': 'gastoHormonios',
        'gastoMaterialOrdenha': 'gastoMaterialOrdenha',
        'gastoEnergiaEletrica': 'gastoEnergiaEletrica',
        'gastoCombustivel': 'gastoCombustivel',
        'gastoReproducao': 'gastoReproducao',
        'gastoImpostoTaxas': 'gastoImpostoTaxas',
        'gastoAdministrativo': 'gastoAdministrativo',
        'gastoReparosConsertos': 'gastoReparosConsertos',
        'gastoAssistenciaTecnica': 'gastoAssistenciaTecnica',
        'gastoCamaAreia': 'gastoCamaAreia',
        'gastoAcessoriosDespesasGerais': 'gastoAcessoriosDespesasGerais',
        'gastoArrendamento': 'gastoArrendamento',
        'consumoLeiteBezerro': 'consumoLeiteBezerro',
        'consumoMaoObraFamiliar': 'consumoMaoObraFamiliar',
        'consumoMaoObraContratada': 'consumoMaoObraContratada',
        'consumoLeiteDescartado': 'consumoLeiteDescartado',
        'observacaoSucedaneo': 'observacaoSucedaneo',
        'observacaoMedicamentosVacinas': 'observacaoMedicamentosVacinas',
        'observacaoMaterialOrdenha': 'observacaoMaterialOrdenha',
        'observacaoReproducao': 'observacaoReproducao',
        'observacaoImpostoTaxas': 'observacaoImpostoTaxas',
        'observacaoHormonios': 'observacaoHormonios',
        'observacaoAssistenciaTecnica': 'observacaoAssistenciaTecnica',
        'observacaoArrendamento': 'observacaoArrendamento',
        'observacaoReparosConsertos': 'observacaoReparosConsertos',
        'observacaoAdministrativo': 'observacaoAdministrativo',
        'observacaoAcessoriosDespesasGerais': 'observacaoAcessoriosDespesasGerais',
        'observacaoCamaAreia': 'observacaoCamaAreia',
        'observacaoLeiteBezerro': 'observacaoLeiteBezerro',
        'observacaoMaoObraFamiliar': 'observacaoMaoObraFamiliar',
        'observacaoMaoObraContratada': 'observacaoMaoObraContratada',
        'observacaoLeiteDescartado': 'observacaoLeiteDescartado',
        'compra_Animais': 'compra_Animais',
        'compraAnimaisObs': 'compraAnimaisObs',
        'compra_Terras': 'compra_Terras',
        'compraTerrasObs': 'compraTerrasObs',
        'parcelaReposicaoCama': 'parcelaReposicaoCama',
        'parcelaCompraAnimais': 'parcelaCompraAnimais',
        'parcelaCompraTerra': 'parcelaCompraTerra',
        'mesReferencia': 'mesReferencia',
        'observacaoCompradeAnimais': 'observacaoCompradeAnimais',
        'observacaoCompradeTerra': 'observacaoCompradeTerra',
        'parcelaSucedaneo': 'parcelaSucedaneo',
        'parcelaMaterialOrdenha': 'parcelaMaterialOrdenha',
        'parcelaReproducao': 'parcelaReproducao',
        'parcelaHormonios': 'parcelaHormonios',
        'parcelaMedicamentosVacinas': 'parcelaMedicamentosVacinas',
        'parcelaImpostoTaxas': 'parcelaImpostoTaxas',
        'parcelaAssistenciaTecnica': 'parcelaAssistenciaTecnica',
        'parcelaArrendamento': 'parcelaArrendamento',
        'parcelaReparosConsertos': 'parcelaReparosConsertos',
        'parcelaAdministrativo': 'parcelaAdministrativo',
        'parcelaAcessoriosDespesasGerais': 'parcelaAcessoriosDespesasGerais',
        'gastoEmprestimosJurosPagos': 'gastoEmprestimosJurosPagos',
        'observacaoEmprestimosJurosPagos': 'observacaoEmprestimosJurosPagos',
        'parcelaEmprestimosJurosPagos': 'parcelaEmprestimosJurosPagos',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'gastoSucedaneo', 'gastoMedicamentosVacinas', 
        'gastoHormonios', 'gastoMaterialOrdenha', 'gastoEnergiaEletrica', 
        'gastoCombustivel', 'gastoReproducao', 'gastoImpostoTaxas', 
        'gastoAdministrativo', 'gastoReparosConsertos', 'gastoAssistenciaTecnica', 
        'gastoCamaAreia', 'gastoAcessoriosDespesasGerais', 'gastoArrendamento', 
        'consumoLeiteBezerro', 'consumoMaoObraFamiliar', 'consumoMaoObraContratada', 
        'consumoLeiteDescartado', 'observacaoSucedaneo', 'observacaoMedicamentosVacinas', 
        'observacaoMaterialOrdenha', 'observacaoReproducao', 'observacaoImpostoTaxas', 
        'observacaoHormonios', 'observacaoAssistenciaTecnica', 'observacaoArrendamento', 
        'observacaoReparosConsertos', 'observacaoAdministrativo', 
        'observacaoAcessoriosDespesasGerais', 'observacaoCamaAreia', 
        'observacaoLeiteBezerro', 'observacaoMaoObraFamiliar', 
        'observacaoMaoObraContratada', 'observacaoLeiteDescartado', 
        'compra_Animais', 'compraAnimaisObs', 'compra_Terras', 'compraTerrasObs', 
        'parcelaReposicaoCama', 'parcelaCompraAnimais', 'parcelaCompraTerra', 
        'mesReferencia', 'observacaoCompradeAnimais', 'observacaoCompradeTerra', 
        'parcelaSucedaneo', 'parcelaMaterialOrdenha', 'parcelaReproducao', 
        'parcelaHormonios', 'parcelaMedicamentosVacinas', 'parcelaImpostoTaxas', 
        'parcelaAssistenciaTecnica', 'parcelaArrendamento', 'parcelaReparosConsertos', 
        'parcelaAdministrativo', 'parcelaAcessoriosDespesasGerais', 
        'gastoEmprestimosJurosPagos', 'observacaoEmprestimosJurosPagos', 
        'parcelaEmprestimosJurosPagos', 'Criado', 'Modificado', 
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'componentescusto')

def alimentacao_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_alimentacao em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'mesReferencia': 'mesReferencia',
        'categoria': 'categoria',
        'item': 'item',
        'operacao': 'operacao',
        'quantidadeComprada': 'quantidadeComprada',
        'quantidadeConsumida': 'quantidadeConsumida',
        'valorUnitario': 'valorUnitario',
        'Parcelamento': 'Parcelamento',
        'Status': 'Status',
        'Collection': 'Collection',
        'Excluido': 'Excluido',
        'Update': 'Update',
        'observacao': 'observacao',
        'id_item': 'id_item',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'mesReferencia', 'categoria', 'item', 'operacao',
        'quantidadeComprada', 'quantidadeConsumida', 'valorUnitario', 
        'Parcelamento', 'Status', 'Collection', 'Excluido', 'Update', 
        'observacao', 'id_item', 'Criado', 'Modificado', 
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'alimentacao')

def area_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_area em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'nomeArea': 'nomeArea',
        'tamanhoArea': 'tamanhoArea',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'precoTerraNua': 'precoTerraNua',
        'Excluido': 'Excluido',
        'Collection': 'Collection',
        'Status': 'Status',
        'observacao': 'observacao',
        'especificacaoArea': 'especificacaoArea',
        'Update': 'Update',
        'arrendamento': 'arrendamento',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'nomeArea', 'tamanhoArea', 'dataInicio', 'dataFim',
        'precoTerraNua', 'Excluido', 'Collection', 'Status', 'observacao',
        'especificacaoArea', 'Update', 'arrendamento', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'area')

def consultor_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_consultor em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idConsultoria': 'idConsultoria',
        'nomeConsultor': 'nomeConsultor',
        'emailConsultor': 'emailConsultor',
        'dteNascConsultor': 'dteNascConsultor',
        'celConsultor': 'celConsultor',
        'generoConsultor': 'generoConsultor',
        'formacaoConsultor': 'formacaoConsultor',
        'cepConsultor': 'cepConsultor',
        'cidadeConsultor': 'cidadeConsultor',
        'ufConsultor': 'ufConsultor',
        'enderecoConsultor': 'enderecoConsultor',
        'preferenciaContato': 'preferenciaContato',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'Status': 'Status',
        'comentarioConsultor': 'comentarioConsultor',
        'Excluido': 'Excluido',
        'aprovacao': 'aprovacao',
        'motivoReprovacao': 'motivoReprovacao',
        'Numero_Residencia_Consultor': 'Numero_Residencia_Consultor',
        'complemento_endereco_consultor': 'complemento_endereco_consultor',
        'idGestor': 'idGestor',
        'cpf_Consultor': 'cpf_Consultor',
        'bairro': 'bairro',
        'logradouro': 'logradouro',
        'editadoPor': 'editadoPor',
        'criadoPor': 'criadoPor',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idConsultoria', 'nomeConsultor', 'emailConsultor', 
        'dteNascConsultor', 'celConsultor', 'generoConsultor', 
        'formacaoConsultor', 'cepConsultor', 'cidadeConsultor', 
        'ufConsultor', 'enderecoConsultor', 'preferenciaContato', 
        'dataInicio', 'dataFim', 'Status', 'comentarioConsultor', 
        'Excluido', 'aprovacao', 'motivoReprovacao', 
        'Numero_Residencia_Consultor', 'complemento_endereco_consultor', 
        'idGestor', 'cpf_Consultor', 'bairro', 'logradouro', 
        'editadoPor', 'criadoPor', 'Criado', 'Modificado', 
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'consultor')

def culturas_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_culturas em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'idArea': 'idArea',
        'nomeCultura': 'nomeCultura',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'Status': 'Status',
        'Collection': 'Collection',
        'Excluido': 'Excluido',
        'Plantio': 'Plantio',
        'observacao': 'observacao',
        'areaha': 'areaha',
        'especificacaoCultura': 'especificacaoCultura',
        'Update': 'Update',
        'identificacaoCultura': 'identificacaoCultura',
        'usoArea': 'usoArea',
        'finalizado': 'finalizado',
        'idEspecificacaoCultura': 'idEspecificacaoCultura',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'idArea', 'nomeCultura', 'dataInicio', 'dataFim',
        'Status', 'Collection', 'Excluido', 'Plantio', 'observacao', 'areaha',
        'especificacaoCultura', 'Update', 'identificacaoCultura', 'usoArea',
        'finalizado', 'idEspecificacaoCultura', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'culturas')

def custoforrageira_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_custoforrageira em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'idCultura': 'idCultura',
        'especificacaoDespesa': 'especificacaoDespesa',
        'qtdeConsumida': 'qtdeConsumida',
        'qtdeComprada': 'qtdeComprada',
        'valorUnitario': 'valorUnitario',
        'item': 'item',
        'AlimentoCultura': 'AlimentoCultura',
        'n': 'n',
        'p': 'p',
        'k': 'k',
        'fonteN': 'fonteN',
        'numeroEspecies': 'numeroEspecies',
        'Status': 'Status',
        'Collection': 'Collection',
        'Excluido': 'Excluido',
        'mesReferencia': 'mesReferencia',
        'Update': 'Update',
        'Operacao': 'Operacao',
        'Parcelas': 'Parcelas',
        'id_item': 'id_item',
        'etapa': 'etapa',
        'idArea': 'idArea',
        'observacao': 'observacao',
        'culturasPlantadas': 'culturasPlantadas',
        'idCulturasPlantadas': 'idCulturasPlantadas',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'idCultura', 'especificacaoDespesa', 'qtdeConsumida',
        'qtdeComprada', 'valorUnitario', 'item', 'AlimentoCultura', 'n', 'p', 'k',
        'fonteN', 'numeroEspecies', 'Status', 'Collection', 'Excluido',
        'mesReferencia', 'Update', 'Operacao', 'Parcelas', 'id_item', 'etapa',
        'idArea', 'observacao', 'culturasPlantadas', 'idCulturasPlantadas',
        'Criado', 'Modificado', 'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'custoforrageira')

def energiacombustivel_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_energiacombustivel em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'qtdeEnergia': 'qtdeEnergia',
        'valorEnergia': 'valorEnergia',
        'qtdeAlcoolgasolina': 'qtdeAlcoolgasolina',
        'valorAlcoolgasolina': 'valorAlcoolgasolina',
        'qtdeDiesel': 'qtdeDiesel',
        'valorDiesel': 'valorDiesel',
        'idFazenda': 'idFazenda',
        'mesReferencia': 'mesReferencia',
        'observacaoEnergia': 'observacaoEnergia',
        'observacaoAlcoolgasolina': 'observacaoAlcoolgasolina',
        'observacaoDiesel': 'observacaoDiesel',
        'qtdeEnergiaSolar': 'qtdeEnergiaSolar',
        'valorEnergiaSolar': 'valorEnergiaSolar',
        'observacaoEnergiaSolar': 'observacaoEnergiaSolar',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'qtdeEnergia', 'valorEnergia', 'qtdeAlcoolgasolina',
        'valorAlcoolgasolina', 'qtdeDiesel', 'valorDiesel', 'idFazenda',
        'mesReferencia', 'observacaoEnergia', 'observacaoAlcoolgasolina',
        'observacaoDiesel', 'qtdeEnergiaSolar', 'valorEnergiaSolar',
        'observacaoEnergiaSolar', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'energiacombustivel')

def entradaestoque_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_entradaestoque em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'Tela': 'Tela',
        'idRelacionamento': 'idRelacionamento',
        'idItem': 'idItem',
        'idFazenda': 'idFazenda',
        'qntComprada': 'qntComprada',
        'valorUnitario': 'valorUnitario',
        'qntConsumida': 'qntConsumida',
        'Excluido': 'Excluido',
        'mesReferencia': 'mesReferencia',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'Tela', 'idRelacionamento', 'idItem', 'idFazenda',
        'qntComprada', 'valorUnitario', 'qntConsumida', 'Excluido',
        'mesReferencia', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'entradaestoque')

def maodeobra_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_maodeobra em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'mesReferencia': 'mesReferencia',
        'tipoCusto': 'tipoCusto',
        'qtdeCustoMaodeObra': 'qtdeCustoMaodeObra',
        'valorCustoMaodeObra': 'valorCustoMaodeObra',
        'familiarQtd': 'familiarQtd',
        'familiarValortotal': 'familiarValortotal',
        'familiarObservacao': 'familiarObservacao',
        'contratataordenhadorQtd': 'contratataordenhadorQtd',
        'contratataordenhadorValortotal': 'contratataordenhadorValortotal',
        'contratataordenhadorObservacao': 'contratataordenhadorObservacao',
        'ctdServicosgeraisQtd': 'ctdServicosgeraisQtd',
        'ctdServicosgeraisValor': 'ctdServicosgeraisValor',
        'ctdServicosgeraisObservacao': 'ctdServicosgeraisObservacao',
        'ctdFolguistaQtde': 'ctdFolguistaQtde',
        'ctdFolguistaValor': 'ctdFolguistaValor',
        'ctdFolguistaObservacao': 'ctdFolguistaObservacao',
        'ctdTratoristaQtde': 'ctdTratoristaQtde',
        'ctdTratoristaValor': 'ctdTratoristaValor',
        'ctdTratoristaObservacao': 'ctdTratoristaObservacao',
        'encargostrabalhistasValor': 'encargostrabalhistasValor',
        'encargostrabalhistasObservacao': 'encargostrabalhistasObservacao',
        'demaisdespesasValor': 'demaisdespesasValor',
        'demaisdespesasObservacao': 'demaisdespesasObservacao',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'mesReferencia', 'tipoCusto', 'qtdeCustoMaodeObra',
        'valorCustoMaodeObra', 'familiarQtd', 'familiarValortotal',
        'familiarObservacao', 'contratataordenhadorQtd',
        'contratataordenhadorValortotal', 'contratataordenhadorObservacao',
        'ctdServicosgeraisQtd', 'ctdServicosgeraisValor',
        'ctdServicosgeraisObservacao', 'ctdFolguistaQtde', 'ctdFolguistaValor',
        'ctdFolguistaObservacao', 'ctdTratoristaQtde', 'ctdTratoristaValor',
        'ctdTratoristaObservacao', 'encargostrabalhistasValor',
        'encargostrabalhistasObservacao', 'demaisdespesasValor',
        'demaisdespesasObservacao', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'maodeobra')

def producao_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_producao em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idItem': 'idItem',
        'areaProduzida': 'areaProduzida',
        'producao': 'producao',
        'dataColheita': 'dataColheita',
        'idFazenda': 'idFazenda',
        'idCultura': 'idCultura',
        'Excluido': 'Excluido',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idItem', 'areaProduzida', 'producao', 'dataColheita',
        'idFazenda', 'idCultura', 'Excluido', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'producao')

def produtor_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_produtor em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idProdutor': 'idProdutor',
        'nomeProdutor': 'nomeProdutor',
        'emailProdutor': 'emailProdutor',
        'dteNascProdutor': 'dteNascProdutor',
        'celProdutor': 'celProdutor',
        'generoProdutor': 'generoProdutor',
        'cepProdutor': 'cepProdutor',
        'cidadeProdutor': 'cidadeProdutor',
        'ufProdutor': 'ufProdutor',
        'enderecoProdutor': 'enderecoProdutor',
        'preferenciaContato': 'preferenciaContato',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'Status': 'Status',
        'Excluido': 'Excluido',
        'dataretorno': 'dataretorno',
        'inativado': 'inativado',
        'motivoReprovacao': 'motivoReprovacao',
        'aprovacao': 'aprovacao',
        'numeroResidencia': 'numeroResidencia',
        'complementoEndereco': 'complementoEndereco',
        'cpf_Produtor': 'cpf_Produtor',
        'bairro': 'bairro',
        'logradouro': 'logradouro',
        'editadoPor': 'editadoPor',
        'criadoPor': 'criadoPor',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idProdutor', 'nomeProdutor', 'emailProdutor', 'dteNascProdutor',
        'celProdutor', 'generoProdutor', 'cepProdutor', 'cidadeProdutor',
        'ufProdutor', 'enderecoProdutor', 'preferenciaContato', 'dataInicio',
        'dataFim', 'Status', 'Excluido', 'dataretorno', 'inativado',
        'motivoReprovacao', 'aprovacao', 'numeroResidencia', 'complementoEndereco',
        'cpf_Produtor', 'bairro', 'logradouro', 'editadoPor', 'criadoPor',
        'Criado', 'Modificado', 'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'produtor')

def qualidadeleite_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_qualidadeleite em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'CCS': 'CCS',
        'CPP': 'CPP',
        'Gordura': 'Gordura',
        'Proteina': 'Proteina',
        'mesReferencia': 'mesReferencia',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'CCS', 'CPP', 'Gordura', 'Proteina',
        'mesReferencia', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'qualidadeleite')

def rebanho_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_rebanho em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'qtdeVacasEmLactacao': 'qtdeVacasEmLactacao',
        'qtdeVacasSecas': 'qtdeVacasSecas',
        'qtdeAnimaisAleitamento': 'qtdeAnimaisAleitamento',
        'qtdeAnimaisRecria': 'qtdeAnimaisRecria',
        'qtdeOutrasCategorias': 'qtdeOutrasCategorias',
        'vidautilOutrasCategorias': 'vidautilOutrasCategorias',
        'valorunitVacasEmLactacao': 'valorunitVacasEmLactacao',
        'valorunitVacasSecas': 'valorunitVacasSecas',
        'valorunitAnimaisAleitamento': 'valorunitAnimaisAleitamento',
        'valorunitAnimaisRecria': 'valorunitAnimaisRecria',
        'valorunitOutrasCategorias': 'valorunitOutrasCategorias',
        'mesReferencia': 'mesReferencia',
        'qtdeMacho': 'qtdeMacho',
        'valorunitMacho': 'valorunitMacho',
        'compraVacasEmLactacao': 'compraVacasEmLactacao',
        'compraVacasSecas': 'compraVacasSecas',
        'compraAnimaisAleitamento': 'compraAnimaisAleitamento',
        'compraAnimaisRecria': 'compraAnimaisRecria',
        'compraMacho': 'compraMacho',
        'compraOutrasCategorias': 'compraOutrasCategorias',
        'vendaVacasEmLactacao': 'vendaVacasEmLactacao',
        'vendaVacasSecas': 'vendaVacasSecas',
        'vendaAnimaisAleitamento': 'vendaAnimaisAleitamento',
        'vendaAnimaisRecria': 'vendaAnimaisRecria',
        'vendaMacho': 'vendaMacho',
        'vendaOutrasCategorias': 'vendaOutrasCategorias',
        'morteVacasEmLactacao': 'morteVacasEmLactacao',
        'morteVacasSecas': 'morteVacasSecas',
        'morteAnimaisAleitamento': 'morteAnimaisAleitamento',
        'morteAnimaisRecria': 'morteAnimaisRecria',
        'morteMacho': 'morteMacho',
        'morteOutrasCategorias': 'morteOutrasCategorias',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'qtdeVacasEmLactacao', 'qtdeVacasSecas',
        'qtdeAnimaisAleitamento', 'qtdeAnimaisRecria', 'qtdeOutrasCategorias',
        'vidautilOutrasCategorias', 'valorunitVacasEmLactacao', 'valorunitVacasSecas',
        'valorunitAnimaisAleitamento', 'valorunitAnimaisRecria', 'valorunitOutrasCategorias',
        'mesReferencia', 'qtdeMacho', 'valorunitMacho', 'compraVacasEmLactacao',
        'compraVacasSecas', 'compraAnimaisAleitamento', 'compraAnimaisRecria',
        'compraMacho', 'compraOutrasCategorias', 'vendaVacasEmLactacao',
        'vendaVacasSecas', 'vendaAnimaisAleitamento', 'vendaAnimaisRecria',
        'vendaMacho', 'vendaOutrasCategorias', 'morteVacasEmLactacao',
        'morteVacasSecas', 'morteAnimaisAleitamento', 'morteAnimaisRecria',
        'morteMacho', 'morteOutrasCategorias', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'rebanho')

def recria_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_recria em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'dataInicio': 'dataInicio',
        'idFazenda': 'idFazenda',
        'dataFim': 'dataFim',
        'percentualVacasCompostBarn': 'percentualVacasCompostBarn',
        'percentualVacasFreeStall': 'percentualVacasFreeStall',
        'percentualVacasSemConfinado': 'percentualVacasSemConfinado',
        'percentualVacasPasto': 'percentualVacasPasto',
        'percentualVacasConfinadoSemEstru': 'percentualVacasConfinadoSemEstrutura',
        'Status': 'Status',
        'Excluido': 'Excluido',
        'Collection': 'Collection',
        'Update': 'Update',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'dataInicio', 'idFazenda', 'dataFim',
        'percentualVacasCompostBarn', 'percentualVacasFreeStall',
        'percentualVacasSemConfinado', 'percentualVacasPasto',
        'percentualVacasConfinadoSemEstrutura', 'Status', 'Excluido',
        'Collection', 'Update', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'recria')

def saidaestoque_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_saidaestoque em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'Tela': 'Tela',
        'idRelacionamento': 'idRelacionamento',
        'idItem': 'idItem',
        'idFazenda': 'idFazenda',
        'qntConsumida': 'qntConsumida',
        'valorUnitario': 'valorUnitario',
        'idEstoque': 'idEstoque',
        'Excluido': 'Excluido',
        'mesReferencia': 'mesReferencia',
        'observacao': 'observacao',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'Tela', 'idRelacionamento', 'idItem', 'idFazenda',
        'qntConsumida', 'valorUnitario', 'idEstoque', 'Excluido',
        'mesReferencia', 'observacao', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'saidaestoque')

def sistemaproducao_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_sistemaproducao em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'idFazenda': 'idFazenda',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'percentualVacasCompostBarn': 'percentualVacasCompostBarn',
        'percentualVacasFreeStall': 'percentualVacasFreeStall',
        'percentualVacasConfinadoSemEstru': 'percentualVacasConfinadoSemEstrutura',
        'percentualVacasSemConfinado': 'percentualVacasSemConfinado',
        'percentualVacasPasto': 'percentualVacasPasto',
        'Status': 'Status',
        'Excluido': 'Excluido',
        'Collection': 'Collection',
        'Update': 'Update',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'dataInicio', 'dataFim',
        'percentualVacasCompostBarn', 'percentualVacasFreeStall',
        'percentualVacasConfinadoSemEstrutura', 'percentualVacasSemConfinado',
        'percentualVacasPasto', 'Status', 'Excluido',
        'Collection', 'Update', 'Criado', 'Modificado',
        'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'sistemaproducao')

def vinculos_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_vinculos em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'IDFazenda': 'idFazenda',
        'IDConsultor': 'idConsultor',
        'dataAprovacao': 'dataAprovacao',
        'dataReprovacao': 'dataReprovacao',
        'dataInativacao': 'dataInativacao',
        'motivoReprovacao': 'motivoReprovacao',
        'IDProdutor': 'idProdutor',
        'motivoSaida': 'motivoSaida',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }
    
    target_columns = [
        'ID', 'idFazenda', 'idConsultor', 'dataAprovacao', 'dataReprovacao',
        'dataInativacao', 'motivoReprovacao', 'idProdutor', 'motivoSaida',
        'Criado', 'Modificado', 'Criado por', 'Modificado por'
    ]
    
    return process_dataframe(items, column_mapping, target_columns, 'vinculos')

    Copiar

def agroindustria_to_dataframe(items: List[Dict]) -> pd.DataFrame:
    """
    Converte os itens da lista tab_agroindustria em DataFrame
    """
    column_mapping = {
        'id': 'ID',
        'momeAgroindustria': 'nomeAgroindustria',
        'cnpjAgroindustria': 'cnpjAgroindustria',
        'razaosocialAgroindustria': 'razaosocialAgroindustria',
        'cepSede': 'cepSede',
        'cidadeSede': 'cidadeSede',
        'dataInicio': 'dataInicio',
        'dataFim': 'dataFim',
        'Status': 'Status',
        'Excluido': 'Excluido',
        'uf': 'uf',
        'cidade': 'cidade',
        'logradouro': 'logradouro',
        'bairro': 'bairro',
        'numero': 'numero',
        'complemento': 'complemento',
        'editadoPor': 'editadoPor',
        'criadoPor': 'criadoPor',
        'Created': 'Criado',
        'Modified': 'Modificado',
        'AuthorLookupId': 'Criado por',
        'EditorLookupId': 'Modificado por'
    }

    target_columns = [
        'ID', 'nomeAgroindustria', 'cnpjAgroindustria', 'razaosocialAgroindustria',
        'cepSede', 'cidadeSede', 'dataInicio', 'dataFim', 'Status', 'Excluido',
        'uf', 'cidade', 'logradouro', 'bairro', 'numero', 'complemento',
        'editadoPor', 'criadoPor', 'Criado', 'Modificado', 'Criado por', 'Modificado por'
    ]

    return process_dataframe(items, column_mapping, target_columns, 'agroindustria')
    
# ========== FUNÇÕES DE SETUP DO SUPABASE ==========
def get_postgres_type(df: pd.DataFrame, column: str) -> str:
    dtype = df[column].dtype
    
    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "TIMESTAMPTZ"
    elif pd.api.types.is_integer_dtype(dtype):
        return "BIGINT"
    elif pd.api.types.is_float_dtype(dtype):
        return "DOUBLE PRECISION"
    elif pd.api.types.is_bool_dtype(dtype):
        return "BOOLEAN"
    else:
        return "TEXT"

def generate_create_table_sql(df: pd.DataFrame, table_name: str) -> str:
    columns_sql = []
    
    for column in df.columns:
        pg_type = get_postgres_type(df, column)
        
        if column.lower() in ['id', 'ID']:
            columns_sql.append(f'    "{column}" {pg_type} PRIMARY KEY')
        else:
            columns_sql.append(f'    "{column}" {pg_type}')
    
    sql = f"""
        -- ========== CRIAR TABELA {table_name.upper()} ==========
        CREATE TABLE IF NOT EXISTS "{table_name}" (
        {',\n'.join(columns_sql)}
        );

        -- ========== CRIAR ÍNDICES (SE NÃO EXISTIREM) ==========
        CREATE INDEX IF NOT EXISTS idx_{table_name}_modificado ON "{table_name}" ("Modificado");
        CREATE INDEX IF NOT EXISTS idx_{table_name}_criado ON "{table_name}" ("Criado");

        -- ========== HABILITAR RLS ==========
        ALTER TABLE "{table_name}" ENABLE ROW LEVEL SECURITY;

        -- ========== REMOVER POLÍTICAS ANTIGAS (SE EXISTIREM) ==========
        DROP POLICY IF EXISTS "service_role_all_access_{table_name}" ON "{table_name}";
        DROP POLICY IF EXISTS "authenticated_read_{table_name}" ON "{table_name}";

        -- ========== CRIAR POLÍTICAS DE ACESSO ==========
        CREATE POLICY "service_role_all_access_{table_name}"
        ON "{table_name}"
        AS PERMISSIVE
        FOR ALL
        TO service_role
        USING (true)
        WITH CHECK (true);

        CREATE POLICY "authenticated_read_{table_name}"
        ON "{table_name}"
        AS PERMISSIVE
        FOR SELECT
        TO authenticated
        USING (true);

        -- ========== COMENTÁRIOS ==========
        COMMENT ON TABLE "{table_name}" IS 'Tabela {table_name} - Dados do SharePoint';
        """
    return sql

def prepare_dataframe_for_supabase(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara o DataFrame para inserção no Supabase (versão robusta com tratamento de NaT)
    """
    df_prepared = df.copy()
    
    # ========== IDENTIFICAR E LIMPAR COLUNAS DE DATA ==========
    date_columns = df_prepared.select_dtypes(include=['datetime64']).columns
    
    for col in date_columns:
        # 1. Substituir NaT por None ANTES de converter
        df_prepared[col] = df_prepared[col].replace({pd.NaT: None})
        
        # 2. Converter apenas valores válidos para ISO 8601
        df_prepared[col] = df_prepared[col].apply(
            lambda x: x.isoformat() if x is not None and pd.notna(x) else None
        )
    
    # ========== SUBSTITUIR QUALQUER NaN/NaT RESTANTE ==========
    df_prepared = df_prepared.replace({pd.NaT: None, float('nan'): None})
    
    # ========== CONVERTER TIPOS NUMÉRICOS ==========
    int_columns = df_prepared.select_dtypes(include=['int64']).columns
    for col in int_columns:
        df_prepared[col] = df_prepared[col].apply(
            lambda x: int(x) if pd.notna(x) else None
        )
    
    float_columns = df_prepared.select_dtypes(include=['float64']).columns
    for col in float_columns:
        df_prepared[col] = df_prepared[col].apply(
            lambda x: float(x) if pd.notna(x) and not pd.isna(x) else None
        )
    
    return df_prepared

print('Funções carregadas!')